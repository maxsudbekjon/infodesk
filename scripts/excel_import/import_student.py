from __future__ import annotations

import re
from collections import Counter
from datetime import date, datetime, time

from django.db.models import Q
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from common import (
    PROJECT_ROOT,
    bootstrap_django,
    clean_text,
    default_password,
    ensure_excel_path,
    extract_phone_number,
    parse_args,
)


bootstrap_django()

from apps.group.models import Group, GroupScore
from apps.pupil.coin import (
    IMPORT_SCORE_REASON,
    calculate_student_coin_offset,
    recalculate_student_total_coin,
)
from apps.pupil.models import Student
from apps.teacher.models import Teacher
from apps.user.models import User


DEFAULT_STUDENT_EXCEL_PATH = PROJECT_ROOT / "CRM-2.xlsx"
CURRENT_YEAR_FOR_AGE = 2026
def load_student_sheets(excel_path: str) -> list[dict]:
    workbook = load_workbook(ensure_excel_path(excel_path), read_only=True, data_only=True)
    sheets = []

    for index, sheet_name in enumerate(workbook.sheetnames, start=1):
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        title = " ".join(clean_text(value) for value in rows[0] if clean_text(value))
        headers = rows[1]
        data_rows = []
        for excel_row_number, row in enumerate(rows[2:], start=3):
            if not any(value not in (None, "") for value in row):
                continue
            data_rows.append({"__excel_row__": excel_row_number, "values": row})

        sheets.append(
            {
                "index": index,
                "sheet_name": sheet_name,
                "title": clean_text(title) or "",
                "headers": headers,
                "rows": data_rows,
            }
        )

    workbook.close()
    return sheets


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_key(value: str | None) -> str:
    if not value:
        return ""
    lowered = normalize_spaces(value).lower()
    lowered = lowered.replace("o'", "o").replace("g'", "g").replace("sh", "s")
    lowered = lowered.replace("’", "").replace("'", "")
    return re.sub(r"[^a-z0-9]+", "", lowered)


def names_match(left: str | None, right: str | None) -> bool:
    left_key = normalize_key(left)
    right_key = normalize_key(right)
    if left_key and right_key and left_key == right_key:
        return True

    left_tokens = [normalize_key(token) for token in normalize_spaces(left or "").split() if normalize_key(token)]
    right_tokens = [normalize_key(token) for token in normalize_spaces(right or "").split() if normalize_key(token)]
    if not left_tokens or not right_tokens:
        return False

    return Counter(left_tokens) == Counter(right_tokens)


def extract_header_phone(header_text: str) -> str | None:
    matches = re.findall(r"(?:\d[\s-]*){9,12}", header_text or "")
    for match in matches:
        phone = extract_phone_number(match)
        if phone:
            return phone
    return None


def teacher_fallback_by_sheet_order(sheet_index: int) -> Teacher | None:
    teachers = list(Teacher.objects.select_related("user").order_by("id"))
    if 1 <= sheet_index <= len(teachers):
        return teachers[sheet_index - 1]
    return None


def find_teacher(sheet_name: str, header_text: str, sheet_index: int) -> Teacher:
    phone = extract_header_phone(header_text)
    if phone:
        teacher = Teacher.objects.select_related("user").filter(user__phone_number=phone).first()
        if teacher:
            return teacher

    search_text = f"{sheet_name} {header_text}"
    search_key = normalize_key(search_text)
    teachers = list(Teacher.objects.select_related("user").all())
    scored: list[tuple[int, Teacher]] = []

    for teacher in teachers:
        full_name = getattr(teacher.user, "full_name", "") or getattr(teacher.user, "display_name", "")
        teacher_key = normalize_key(full_name)
        score = 0
        if teacher_key and teacher_key in search_key:
            score += 5
        teacher_tokens = [normalize_key(token) for token in normalize_spaces(full_name).split() if len(token) >= 3]
        for token in teacher_tokens:
            if token and token in search_key:
                score += 2

        if score:
            scored.append((score, teacher))

    if scored:
        scored.sort(key=lambda item: (-item[0], item[1].id))
        if len(scored) == 1 or scored[0][0] > scored[1][0]:
            return scored[0][1]

    fallback = teacher_fallback_by_sheet_order(sheet_index)
    if fallback:
        print(
            f"[WARN] Teacher aniq topilmadi, sheet tartibi bo'yicha fallback ishlatildi: "
            f"sheet={sheet_name!r} -> teacher_id={fallback.id}"
        )
        return fallback

    raise ValueError(f"Teacher topilmadi: sheet={sheet_name!r}, header={header_text!r}")


def parse_day_and_time(value: object) -> tuple[str, time]:
    text = normalize_spaces(str(value or "")).replace(";", ":")
    match = re.match(r"^(?P<prefix>[dDsS])\s*/?\s*(?P<time>\d{1,2}:\d{2})$", text)
    if not match:
        raise ValueError(f"VAQT formatini tushunib bo'lmadi: {value}")

    prefix = match.group("prefix").lower()
    if prefix == "d":
        days_choice = "odd_days"
    elif prefix == "s":
        days_choice = "even_days"
    else:
        raise ValueError(f"VAQT formatini tushunib bo'lmadi: {value}")

    parsed_time = datetime.strptime(match.group("time"), "%H:%M").time()
    return days_choice, parsed_time


def detect_column_indexes(headers: object) -> dict[str, object]:
    phone_indexes: list[int] = []
    contract_indexes: list[int] = []
    arrival_indexes: list[int] = []
    indexes: dict[str, object] = {
        "last_name": None,
        "first_name": None,
        "full_name": None,
        "birthday": None,
        "arrival_indexes": arrival_indexes,
        "time": None,
        "coin": None,
        "phone_indexes": phone_indexes,
        "contract_indexes": contract_indexes,
    }

    for index, header in enumerate(headers or []):
        text = normalize_spaces(str(header or "")).lower()
        if (
            "famili" in text
            or "family" in text
            or text.startswith("fam")
            or "фам" in text
            or "last name" in text
        ):
            indexes["last_name"] = index
        elif text in {"ism", "ismi"} or text.endswith(" ism") or "имя" in text or "first name" in text:
            indexes["first_name"] = index
        elif (
            "fio" in text
            or "full name" in text
            or "full_name" in text
            or "ism famili" in text
            or "ism familya" in text
            or "fish" in text
            or "фио" in text
            or "talaba" in text
            or "student" in text
            or "oquvchi" in text
        ):
            indexes["full_name"] = index
        elif "telefon" in text or "tel nomer" in text:
            phone_indexes.append(index)
        elif "birth" in text or "brith" in text:
            indexes["birthday"] = index
        elif "kelgan" in text or "sana" in text:
            arrival_indexes.append(index)
        elif "vaqt" in text:
            indexes["time"] = index
        elif "coin" in text:
            indexes["coin"] = index
        elif "shart" in text:
            contract_indexes.append(index)

    birthday_index = indexes["birthday"]
    if phone_indexes and birthday_index is not None:
        for extra_index in range(phone_indexes[-1] + 1, birthday_index):
            phone_indexes.append(extra_index)

    return indexes


def row_value(values: list[object], index: int | None):
    if index is None or index >= len(values):
        return None
    return values[index]


def time_distance_minutes(left: time, right: time) -> int:
    left_minutes = left.hour * 60 + left.minute
    right_minutes = right.hour * 60 + right.minute
    return abs(left_minutes - right_minutes)


def nearest_group_for_teacher(teacher: Teacher, days_choice: str, start_lesson: time) -> Group | None:
    candidates = list(
        Group.objects.select_related("course", "teacher__user", "branch")
        .filter(teacher=teacher, lessons_days_choice=days_choice)
        .order_by("start_lesson", "id")
    )
    if not candidates:
        return None

    ranked = sorted(
        candidates,
        key=lambda group: (
            time_distance_minutes(group.start_lesson, start_lesson),
            group.start_lesson > start_lesson,
            group.id,
        ),
    )
    nearest = ranked[0]
    if time_distance_minutes(nearest.start_lesson, start_lesson) <= 60:
        return nearest
    return None


def group_candidates_for_teacher(teacher: Teacher, days_choice: str, start_lesson: time) -> list[Group]:
    return list(
        Group.objects.select_related("course", "teacher__user", "branch")
        .filter(teacher=teacher, lessons_days_choice=days_choice, start_lesson=start_lesson)
        .order_by("id")
    )


def select_group(teacher: Teacher, sheet_name: str, header_text: str, days_choice: str, start_lesson: time) -> Group:
    candidates = group_candidates_for_teacher(teacher, days_choice, start_lesson)
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        sheet_key = f"{sheet_name} {header_text}"
        for candidate in candidates:
            candidate_key = normalize_key(
                f"{candidate.title} {candidate.course.name} {getattr(candidate.teacher.user, 'full_name', '')}"
            )
            if candidate_key and any(token and token in candidate_key for token in _interesting_tokens(sheet_key)):
                return candidate
        raise ValueError(
            f"Bir nechta group topildi: teacher_id={teacher.id}, days={days_choice}, start={start_lesson}"
        )

    nearest_teacher_group = nearest_group_for_teacher(teacher, days_choice, start_lesson)
    if nearest_teacher_group:
        print(
            f"[WARN] Exact group topilmadi, eng yaqin teacher group tanlandi: "
            f"sheet={sheet_name!r}, teacher_id={teacher.id}, requested={start_lesson}, "
            f"matched={nearest_teacher_group.start_lesson}, group={nearest_teacher_group.title}"
        )
        return nearest_teacher_group

    fallback_candidates = list(
        Group.objects.select_related("course", "teacher__user", "branch")
        .filter(lessons_days_choice=days_choice, start_lesson=start_lesson)
        .order_by("id")
    )
    sheet_tokens = _interesting_tokens(f"{sheet_name} {header_text}")
    for candidate in fallback_candidates:
        candidate_text = normalize_key(
            f"{candidate.title} {candidate.course.name} {getattr(candidate.teacher.user, 'full_name', '')}"
        )
        if any(token and token in candidate_text for token in sheet_tokens):
            return candidate

    raise ValueError(
        f"Group topilmadi: sheet={sheet_name!r}, teacher_id={teacher.id}, days={days_choice}, start={start_lesson}"
    )


def _interesting_tokens(normalized_text: str) -> list[str]:
    raw_tokens = re.findall(r"[a-z0-9']+", normalize_spaces(str(normalized_text or "")).lower())
    stopwords = {
        "ingliz",
        "tili",
        "foundation",
        "kids",
        "full",
        "web",
        "dasturlash",
        "grafik",
        "dizayn",
        "kiberxavfsizlik",
        "matematika",
        "rus",
        "smm",
        "ai",
        "opa",
    }
    normalized_stopwords = {normalize_key(token) for token in stopwords}
    tokens: list[str] = []
    for token in raw_tokens:
        normalized = normalize_key(token)
        if len(normalized) >= 3 and normalized not in normalized_stopwords:
            tokens.append(normalized)
    return tokens


def normalize_contract(value: object) -> bool:
    text = normalize_key(clean_text(value) or "")
    return text.startswith("bor") or "shartnoma" in text


def normalize_arrival_date(value: object) -> date | None:
    if value is None or normalize_contract(value):
        return None
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        year = 2026 if 1 <= value.month <= 3 else 2025
        return value.replace(year=year)

    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, int):
        return None

    text = clean_text(value)
    if not text:
        return None
    normalized_text = text.replace("/", ".").replace("-", ".")
    match = re.fullmatch(r"(?P<day>\d{1,2})\.(?P<month>\d{1,2})(?:\.(?P<year>\d{2,4}))?", normalized_text)
    if not match:
        return None

    day = int(match.group("day"))
    month = int(match.group("month"))
    if not 1 <= day <= 31 or not 1 <= month <= 12:
        return None
    year = 2026 if 1 <= month <= 3 else 2025
    return date(year, month, day)


def extract_arrival_date(values: list[object], arrival_indexes: list[int]) -> date | None:
    for arrival_index in arrival_indexes:
        normalized = normalize_arrival_date(row_value(values, arrival_index))
        if normalized:
            return normalized
    return None


def normalize_birthday(value: object) -> date | None:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    if isinstance(value, int):
        if not 1 <= value <= 120:
            return None
        return date(CURRENT_YEAR_FOR_AGE - value, 1, 1)

    text = clean_text(value)
    if not text:
        return None
    normalized_text = text.replace("/", ".").replace("-", ".")
    match = re.fullmatch(r"(?P<day>\d{1,2})\.(?P<month>\d{1,2})\.(?P<year>\d{2,4})", normalized_text)
    if match:
        day = int(match.group("day"))
        month = int(match.group("month"))
        year = int(match.group("year"))
        if year < 100:
            year += 2000
        try:
            return date(year, month, day)
        except ValueError:
            return None
    if text.isdigit():
        age = int(text)
        if not 1 <= age <= 120:
            return None
        return date(CURRENT_YEAR_FOR_AGE - age, 1, 1)
    return None


def normalize_coin(value: object) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, float):
        return int(value)
    return int(str(value))


def build_full_name(last_name: object, first_name: object) -> str:
    parts = [clean_text(last_name), clean_text(first_name)]
    return " ".join(part for part in parts if part)


def guess_full_name_from_row(values: list[object]) -> str | None:
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        if extract_phone_number(text):
            continue
        # Skip pure numeric/time-like cells
        if re.fullmatch(r"[\\d\\s\\+\\-\\.\\/()]+", text):
            continue
        if len(text) < 3:
            continue
        return text
    return None


def choose_primary_phone(*phone_values: object) -> tuple[str | None, str | None]:
    phones: list[str] = []
    for raw_value in phone_values:
        phone = extract_phone_number(raw_value)
        if phone and phone not in phones:
            phones.append(phone)

    primary_phone = phones[0] if phones else None
    secondary_phone = phones[1] if len(phones) > 1 else None
    return primary_phone, secondary_phone


def make_aware_midday(value: date | None):
    if value is None:
        return None
    naive = datetime.combine(value, time(12, 0))
    return timezone.make_aware(naive)


def resolve_secondary_phone(user: User | None, primary_phone: str | None, secondary_phone: str | None) -> str | None:
    if not secondary_phone or secondary_phone == primary_phone:
        return None

    phone_busy = User.objects.filter(phone_number=secondary_phone)
    phone2_busy = User.objects.filter(phone_number2=secondary_phone)
    if user:
        phone_busy = phone_busy.exclude(pk=user.pk)
        phone2_busy = phone2_busy.exclude(pk=user.pk)

    if phone_busy.exists() or phone2_busy.exists():
        return None
    return secondary_phone


def can_reuse_user_for_student(user: User, full_name: str) -> bool:
    existing_student = Student.objects.filter(user=user).first()
    if existing_student and existing_student.full_name and not names_match(existing_student.full_name, full_name):
        return False

    if user.full_name and not names_match(user.full_name, full_name):
        if existing_student:
            return False
        if user.role != "student":
            return False
    return True


def upsert_student_user(full_name: str, primary_phone: str | None, secondary_phone: str | None, birthday: date | None):
    if not primary_phone:
        return None

    user = User.objects.filter(phone_number=primary_phone).first()
    if not user:
        safe_secondary_phone = resolve_secondary_phone(None, primary_phone, secondary_phone)
        user = User.objects.create_user(
            phone_number=primary_phone,
            password=default_password(primary_phone),
            full_name=full_name,
            role="student",
            birthday=birthday,
            phone_number2=safe_secondary_phone,
        )
        return user

    if not can_reuse_user_for_student(user, full_name):
        return None

    update_fields = []
    if full_name and user.full_name != full_name:
        user.full_name = full_name
        update_fields.append("full_name")
    if birthday and user.birthday != birthday:
        user.birthday = birthday
        update_fields.append("birthday")
    safe_secondary_phone = resolve_secondary_phone(user, user.phone_number, secondary_phone)
    if safe_secondary_phone and user.phone_number2 != safe_secondary_phone:
        user.phone_number2 = safe_secondary_phone
        update_fields.append("phone_number2")

    if update_fields:
        user.save(update_fields=update_fields)
    return user


def find_existing_student(user: User | None, full_name: str, primary_phone: str | None, group: Group) -> Student | None:
    base_queryset = Student.objects.select_related("user", "group").prefetch_related("groups").filter(
        center=group.course.center
    )

    if user:
        student = Student.objects.filter(user=user).first()
        if student and _student_name_matches(student, full_name):
            return student

    if primary_phone:
        phone_candidates = base_queryset.filter(
            Q(phone_number=primary_phone)
            | Q(user__phone_number=primary_phone)
            | Q(user__phone_number2=primary_phone)
        ).distinct()
        matched_student = _pick_matching_student(phone_candidates, full_name, primary_phone, group)
        if matched_student:
            return matched_student

    if full_name:
        exact_group_candidates = base_queryset.filter(Q(group=group) | Q(groups=group)).distinct()
        matched_student = _pick_matching_student(exact_group_candidates, full_name, primary_phone, group)
        if matched_student:
            return matched_student

        same_course_candidates = base_queryset.filter(
            Q(group__course=group.course) | Q(groups__course=group.course)
        ).distinct()
        matched_student = _pick_matching_student(same_course_candidates, full_name, primary_phone, group)
        if matched_student:
            return matched_student

        matched_student = _pick_matching_student(base_queryset, full_name, primary_phone, group)
        if matched_student:
            return matched_student

    return None


def _student_name_matches(student: Student, full_name: str) -> bool:
    if names_match(student.full_name, full_name):
        return True

    user = getattr(student, "user", None)
    if user and names_match(user.display_name, full_name):
        return True

    return False


def _student_phones(student: Student) -> set[str]:
    phones = {student.phone_number}
    user = getattr(student, "user", None)
    if user:
        phones.add(user.phone_number)
        phones.add(user.phone_number2)
    return {phone for phone in phones if phone}


def _pick_matching_student(candidates, full_name: str, primary_phone: str | None, group: Group) -> Student | None:
    matched_by_name = [student for student in candidates if _student_name_matches(student, full_name)]
    if not matched_by_name:
        return None

    if primary_phone:
        matched_by_phone = [
            student for student in matched_by_name if primary_phone in _student_phones(student)
        ]
        if len(matched_by_phone) == 1:
            return matched_by_phone[0]
        if matched_by_phone:
            matched_by_name = matched_by_phone

    exact_group_matches = [
        student
        for student in matched_by_name
        if student.group_id == group.id or any(related_group.id == group.id for related_group in student.groups.all())
    ]
    if len(exact_group_matches) == 1:
        return exact_group_matches[0]
    if exact_group_matches:
        matched_by_name = exact_group_matches

    same_course_matches = [
        student
        for student in matched_by_name
        if (
            student.group and student.group.course_id == group.course_id
        ) or any(related_group.course_id == group.course_id for related_group in student.groups.all())
    ]
    if len(same_course_matches) == 1:
        return same_course_matches[0]
    if same_course_matches:
        matched_by_name = same_course_matches

    if len(matched_by_name) == 1:
        return matched_by_name[0]

    return None


def get_or_create_student(full_name: str, primary_phone: str | None, group: Group, contract: bool, user: User | None, coin: int):
    student = find_existing_student(user, full_name, primary_phone, group)

    created = False
    if not student:
        student = Student.objects.create(
            user=user,
            full_name=full_name,
            phone_number=primary_phone,
            group=group,
            center=group.course.center,
            contract=contract,
            used_coin=0,
        )
        created = True
    else:
        update_fields = []
        if user and student.user_id != user.id:
            student.user = user
            update_fields.append("user")
        if full_name and student.full_name != full_name:
            student.full_name = full_name
            update_fields.append("full_name")
        if primary_phone and student.phone_number != primary_phone:
            student.phone_number = primary_phone
            update_fields.append("phone_number")
        if student.center_id != group.course.center_id:
            student.center = group.course.center
            update_fields.append("center")
        if student.group_id is None:
            student.group = group
            update_fields.append("group")
        elif student.group_id != group.id and student.group and student.group.course_id == group.course_id:
            student.group = group
            update_fields.append("group")
        if student.contract != contract:
            student.contract = contract
            update_fields.append("contract")
        if update_fields:
            student.save(update_fields=update_fields)

    _sync_import_group_membership(student, group)
    _sync_import_coin(student, group, coin)
    return student, created


def _sync_import_group_membership(student: Student, group: Group) -> None:
    stale_scores = GroupScore.objects.filter(
        student=student,
        reason=IMPORT_SCORE_REASON,
        group__course=group.course,
    ).exclude(group=group)
    stale_group_ids = list(stale_scores.values_list("group_id", flat=True))
    if stale_group_ids:
        stale_scores.delete()
        student.groups.remove(*stale_group_ids)

    same_course_group_ids = list(
        student.groups.filter(course=group.course).exclude(pk=group.pk).values_list("pk", flat=True)
    )
    removable_group_ids = [group_id for group_id in same_course_group_ids if group_id not in stale_group_ids]
    if removable_group_ids:
        student.groups.remove(*removable_group_ids)

    student.groups.add(group)
    if student.group_id is None or (student.group_id != group.id and student.group and student.group.course_id == group.course_id):
        Student.objects.filter(pk=student.pk).update(group=group)
        student.group = group


def _sync_import_coin(student: Student, _group: Group, coin: int) -> None:
    GroupScore.objects.filter(student=student, reason=IMPORT_SCORE_REASON).delete()

    target_offset = calculate_student_coin_offset(student.id, coin)
    if student.coin_offset != target_offset:
        Student.objects.filter(pk=student.pk).update(coin_offset=target_offset)
        student.coin_offset = target_offset

    recalculate_student_total_coin(student.id)


def import_students(excel_path: str) -> None:
    sheets = load_student_sheets(excel_path)
    created_count = 0
    updated_count = 0
    skipped_count = 0

    for sheet in sheets:
        try:
            teacher = find_teacher(sheet["sheet_name"], sheet["title"], sheet["index"])
            column_indexes = detect_column_indexes(sheet["headers"])
            print(f"[SHEET] {sheet['sheet_name']} -> teacher_id={teacher.id}")
        except Exception as exc:
            skipped_count += len(sheet["rows"])
            print(
                f"[SKIPPED SHEET] {sheet['sheet_name']} teacher yoki header muammosi sabab o'tkazib yuborildi: {exc}"
            )
            continue

        for row in sheet["rows"]:
            try:
                with transaction.atomic():
                    values = list(row["values"])
                    raw_full_name = row_value(values, column_indexes.get("full_name"))
                    if clean_text(raw_full_name):
                        full_name = clean_text(raw_full_name)
                    else:
                        full_name = build_full_name(
                            row_value(values, column_indexes["last_name"]),
                            row_value(values, column_indexes["first_name"]),
                        )
                    if not full_name:
                        full_name = guess_full_name_from_row(values)
                    if not full_name:
                        skipped_count += 1
                        print(f"[SKIPPED] {sheet['sheet_name']} row={row['__excel_row__']} full_name bo'sh")
                        continue

                    phone_indexes = column_indexes["phone_indexes"]
                    contract_indexes = column_indexes["contract_indexes"]
                    arrival_indexes = column_indexes["arrival_indexes"]

                    primary_phone, secondary_phone = choose_primary_phone(
                        *[row_value(values, phone_index) for phone_index in phone_indexes]
                    )
                    birthday = normalize_birthday(row_value(values, column_indexes["birthday"]))
                    contract = any(
                        normalize_contract(row_value(values, contract_index)) for contract_index in contract_indexes
                    )
                    if not contract:
                        contract = any(
                            normalize_contract(row_value(values, arrival_index)) for arrival_index in arrival_indexes
                        )
                    arrival_date = extract_arrival_date(values, arrival_indexes)
                    raw_time = row_value(values, column_indexes["time"])
                    if clean_text(raw_time) is None:
                        skipped_count += 1
                        print(
                            f"[SKIPPED] {sheet['sheet_name']} row={row['__excel_row__']} "
                            f"DARS VAQTI bo'sh"
                        )
                        continue
                    try:
                        days_choice, start_lesson = parse_day_and_time(raw_time)
                    except ValueError as exc:
                        skipped_count += 1
                        print(
                            f"[SKIPPED] {sheet['sheet_name']} row={row['__excel_row__']} "
                            f"{exc}"
                        )
                        continue
                    coin = normalize_coin(row_value(values, column_indexes["coin"]))

                    try:
                        group = select_group(teacher, sheet["sheet_name"], sheet["title"], days_choice, start_lesson)
                    except ValueError as exc:
                        skipped_count += 1
                        print(
                            f"[SKIPPED] {sheet['sheet_name']} row={row['__excel_row__']} "
                            f"{exc}"
                        )
                        continue
                    user = upsert_student_user(full_name, primary_phone, secondary_phone, birthday)
                    student, created = get_or_create_student(full_name, primary_phone, group, contract, user, coin)

                    if arrival_date:
                        aware_dt = make_aware_midday(arrival_date)
                        Student.objects.filter(pk=student.pk).update(created_at=aware_dt)

                    if created:
                        created_count += 1
                        print(
                            f"[CREATED] student row={row['__excel_row__']} "
                            f"name={full_name} group={group.title}"
                        )
                    else:
                        updated_count += 1
                        print(
                            f"[UPDATED] student row={row['__excel_row__']} "
                            f"name={full_name} group={group.title}"
                        )
            except Exception as exc:
                skipped_count += 1
                print(
                    f"[SKIPPED] {sheet['sheet_name']} row={row['__excel_row__']} "
                    f"kutilmagan xato: {exc}"
                )
                continue
    print(f"Student import yakunlandi. created={created_count}, updated={updated_count}, skipped={skipped_count}")


if __name__ == "__main__":
    args = parse_args("Import students from CRM.xlsx", DEFAULT_STUDENT_EXCEL_PATH)
    excel_path = args.excel_path or DEFAULT_STUDENT_EXCEL_PATH
    import_students(excel_path)
