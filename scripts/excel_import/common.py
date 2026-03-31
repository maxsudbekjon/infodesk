from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_EXCEL_PATH = Path("~/Desktop/oquv_markaz_template (2) (2).xlsx").expanduser()
PROJECT_ROOT = Path(__file__).resolve().parents[2]


def bootstrap_django() -> None:
    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))

    os.environ.setdefault("DJANGO_ENV", "dev")
    os.environ.setdefault("DEBUG", "false")

    django_env = os.environ.get("DJANGO_ENV", "dev").lower()
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", f"config.settings.{django_env}")

    import django

    django.setup()


def parse_args(description: str, default_excel_path: str | Path | None = None) -> argparse.Namespace:
    default_path = str(default_excel_path or DEFAULT_EXCEL_PATH)
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument(
        "excel_path",
        nargs="?",
        default=default_path,
        help=f"Excel file path. Default: {default_path}",
    )
    return parser.parse_args()


def ensure_excel_path(excel_path: str | Path) -> Path:
    path = Path(excel_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Excel file topilmadi: {path}")
    return path


def normalize_header(value: object, index: int) -> str:
    if value is None:
        return f"column_{index}"
    header = re.sub(r"\s+", " ", str(value)).strip().lower()
    return header or f"column_{index}"


def has_value(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def load_sheet_rows(excel_path: str | Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(ensure_excel_path(excel_path), read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(value, index + 1) for index, value in enumerate(rows[0])]
    payload = []

    for row_number, row in enumerate(rows[1:], start=1):
        if not any(has_value(value) for value in row):
            continue

        item = {"__row__": row_number}
        for index, header in enumerate(headers):
            item[header] = row[index] if index < len(row) else None
        payload.append(item)

    workbook.close()
    return payload


def row_map(rows: list[dict]) -> dict[int, dict]:
    return {row["__row__"]: row for row in rows}


def require_row(rows_by_number: dict[int, dict], row_number: int, sheet_name: str) -> dict:
    row = rows_by_number.get(row_number)
    if not row:
        raise ValueError(f"`{sheet_name}` sheetida {row_number}-qator topilmadi.")
    return row


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def _value_to_digit_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\D", "", str(value))


def normalize_phone_number(value: object) -> str | None:
    text = clean_text(value)
    if not text:
        return None

    digits = _value_to_digit_text(value)
    if len(digits) == 9:
        digits = f"998{digits}"
    if len(digits) < 12:
        raise ValueError(f"Telefon raqamni normallashtirib bo'lmadi: {value}")
    return f"+{digits}"


def extract_phone_number(value: object) -> str | None:
    digits = _value_to_digit_text(value)
    if not digits:
        return None
    if len(digits) == 9:
        return f"+998{digits}"
    if len(digits) == 12 and digits.startswith("998"):
        return f"+{digits}"
    return None


def default_password(phone_number: str) -> str:
    return phone_number.replace("+", "")


def normalize_decimal(value: object, default: Decimal | None = None) -> Decimal | None:
    if value is None or value == "":
        return default
    return Decimal(str(value).strip())


def normalize_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.fromisoformat(str(value)).date()


def normalize_time(value: object) -> time | None:
    if value is None or value == "":
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    return time.fromisoformat(str(value))


def parse_row_references(value: object) -> list[int]:
    if value is None or value == "":
        return []
    if isinstance(value, int):
        return [value]
    if isinstance(value, float):
        return [int(value)]

    digits = re.findall(r"\d+", str(value))
    return [int(item) for item in digits]


def first_row_reference(value: object) -> int | None:
    refs = parse_row_references(value)
    return refs[0] if refs else None


def normalize_owner_role(value: object) -> str:
    text = (clean_text(value) or "ceo").lower()
    mapping = {
        "seo": "ceo",
        "ceo": "ceo",
        "owner": "ceo",
        "admin": "admin",
        "teacher": "teacher",
        "student": "student",
        "user": "user",
        "manager": "meneger",
        "meneger": "meneger",
    }
    return mapping.get(text, "ceo")


def normalize_days_choice(value: object) -> str:
    text = (clean_text(value) or "").lower()
    mapping = {
        "toq": "odd_days",
        "odd": "odd_days",
        "juft": "even_days",
        "even": "even_days",
        "har kuni": "every_day",
        "every day": "every_day",
        "every_day": "every_day",
    }
    normalized = mapping.get(text)
    if not normalized:
        raise ValueError(f"Dars kunlari qiymati tushunarsiz: {value}")
    return normalized
